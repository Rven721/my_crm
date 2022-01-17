import os
import django
from openpyxl import Workbook

os.environ['DJANGO_SETTINGS_MODULE'] = 'conf.settings.base'
django.setup()
from crm.models import Project

event_type_decoder = {
    'text_consult': 'Письменная консультация',
    'phone_consult': 'Телефонная консультация',
    'call': 'ВКС',
    'document_transfer': 'Обмен документами',
    'paper_check': 'Работа с документами',
    'data_request': 'Запрос информации',
    'project_update': 'Изменение статуса',
    'discuss': 'Обсуждение',
    'due_dil': 'Контрактация',
}

project_status_decoder = {
    'new': 'Получена анкета. Ожидается финальное обсуждение и согласование',
    'progress': 'Идет работа по подготовке заявки',
    'finished': 'Работа по подготовке заявки завершена'
}


def project_event_history_report(project_id):
    """Принимает на вход id проекта и возвращает несохраненный excel файл
        c историей смены статусов по проекту"""
    project = Project.objects.get(id=project_id)
    event_history = project.events.all()
    wb = Workbook()
    ws = wb.active
    ws['a1'] = 'Проект:'
    ws['a2'] = 'Текущий статус'
    ws['b1'] = project.name
    ws['b2'] = project_status_decoder[project.statuses.all().last().status]
    ws.append([''])
    ws.append(['Тип', 'Дата', 'Описание', 'Результат'])
    for event in event_history:
        data = [event_type_decoder[event.type], event.date, event.description, event.result]
        ws.append(data)
    return wb


def project_status_report(company=None, project_deliver=None, project_status=None):
    """Функция возвращает несохраненный excel файл содержащий список всех проектов и описание их текущего статуса для
    заданной компании от заданного источника"""
    header = ['Компания', 'Проект', 'Текущий статус', 'Последнее событие', 'Источник']
    project_list = Project.objects.all()
    if project_deliver:
        project_list = project_list.filter(project_deliver=project_deliver)
    if company:
        project_list = project_list.filter(company=company)
    if project_status is not None:
        project_list = list(project_list)
        check_list = project_list.copy()
        for project in check_list:
            try:
                if project.statuses.last().status != project_status:
                    project_list.remove(project)
            except AttributeError:
                project_list.remove(project)
    wb = Workbook()
    ws = wb.active
    ws.append(header)
    for project in project_list:
        try:
            last_event_result = project.events.last().result
        except AttributeError:
            last_event_result = 'Данные отсутствуют'
        project_data = [
            project.company.short_name,
            project.full_name,
            project_status_decoder[project.statuses.all().last().status],
            last_event_result, project.project_deliver.name
        ]
        ws.append(project_data)
    return wb
