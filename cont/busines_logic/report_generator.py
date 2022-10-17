"""Report generator for contracts"""
from openpyxl import Workbook

REPORT_COLUMNS = {
    'rep_number': 'Номер',
    'rep_date': 'Дата',
    'rep_project': 'Проект',
    'rep_agent': 'Агент',
    'rep_terms': 'Тариф',
    'rep_status': 'Статус',
    'rep_comment': 'Комментарий',
    'rep_fix_1': 'Фикс_1',
    'rep_fix_2': 'Фикс_2',
    'rep_fix_total': 'Фикс_общий',
    'rep_success_fee': 'Плата за успех',
    'rep_team_fee': 'Вознаграждение команды',
}


def get_columns(request):
    """Will return a lis of columns for report according to request params"""
    req_colums = ['rep' in param for param in request.GET]
    columns = [value for key, value in REPORT_COLUMNS.items() if key in req_colums]
    return columns


def extract_data(contract, columns):
    """Will return contract data for given columns"""
    contract_data = {
        'rep_number': contract.number,
        'rep_date': contract.date,
        'rep_project': contract.project,
        'rep_agent': contract.agent,
        'rep_terms': contract.terms,
        'rep_status': contract.payment_status,
        'rep_comment': contract.comment,
        'rep_fix_1': contract.get_revenue_fix_pt1,
        'rep_fix_2': contract.get_revenue_fix_pt2,
        'rep_fix_total': contract.get_revenue_fix,
        'rep_success_fee': contract.get_success_fee,
        'rep_team_fee': contract.get_team_revenue,
    }
    result = [value for field, value in contract_data.items() if field in columns]
    return result


def get_contracts_columns_report(contracts, columns=None):
    """Will return an excel file according to the request"""
    if columns:
        req_colums = ['rep' in param for param in columns]
    else:
        req_colums = list(REPORT_COLUMNS.values())
    contracts_for_report = [extract_data(contract, req_colums) for contract in contracts]
    header = [value for key, value in REPORT_COLUMNS.items() if key in req_colums]
    wb = Workbook()
    ws = wb.active
    ws.append(header)
    ws.merge_cells('B1:F1')
    ws.merge_cells('B2:F2')
    for contract in contracts_for_report:
        ws.append(contract)
    return wb
