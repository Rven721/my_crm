"""Module to extract data for smeta"""
from openpyxl import load_workbook


def get_workers(file_name: str) -> list:
    """will extract a list of workers from prepared excel file"""
    workbook = load_workbook(filename=file_name)
    workbook_ws = workbook['workers']
    workers = []
    for row in workbook_ws:
        worker = tuple(cell.value for cell in row)
        workers.append(worker)
    return workers

def get_events(file_name: str) -> list:
    """Will return a list of vents from prepared ecxel file"""
    workbook = load_workbook(filename=file_name)
    workbook_ws = workbook['events']
    events = []
    for row in workbook_ws:
        event = [cell.value for cell in row]
        event[1] = tuple(int(i) for i in event[1].split(','))
        event[2] = tuple(int(i) for i in event[2].split(','))
        events.append(tuple(event))
    return events

def get_monthes(file_name: str) -> list:
    """Will return a list of monthes from prepared ecxel file"""
    workbook = load_workbook(filename=file_name)
    workbook_ws = workbook['monthes']
    monthes = []
    for row in workbook_ws:
        raw_month = [cell.value for cell in row]
        date = (raw_month[1].month, raw_month[1].year)
        work_hours = raw_month[2]
        stage = raw_month[0]
        month = (date, work_hours, stage)
        monthes.append(month)
    return monthes
